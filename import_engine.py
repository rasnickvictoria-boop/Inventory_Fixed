from __future__ import annotations

import re
import tempfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests
from openpyxl import load_workbook
from pypdf import PdfReader


SECTION_NAMES = [
    "Consumables & Replenishment",
    "Kitchen Inventory",
    "Technology & Guest Experience",
    "Bedding & Linens",
    "Bathrooms",
    "Safety",
    "Misc. Items",
    "Miscellaneous",
    "Property Condition Assessment",
    "Action Items",
]

DATE_RE = re.compile(
    r"(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),\s+"
    r"([A-Za-z]+)\s+(\d{1,2}),\s+(\d{4})"
)


@dataclass
class ParsedReport:
    property_name: str
    inspection_date: datetime
    address: str
    source: str
    inspection_id: str
    inventory: list[list] = field(default_factory=list)
    conditions: list[list] = field(default_factory=list)
    feedback: list[list] = field(default_factory=list)
    ratings: list[list] = field(default_factory=list)


def clean_text(value: str) -> str:
    replacements = {
        "\uf0a2": "",
        "\uf128": "",
        "\uf005": "★",
        "\uf006": "☆",
        "\uf0c5": "",
        "\uf03d": "",
        "\uf00c": "",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)

    value = re.sub(r"\bT\s+ext\b", "Text", value, flags=re.I)
    value = re.sub(r"\bY\s+es\b", "Yes", value, flags=re.I)
    value = re.sub(r"\bT\s+owel\b", "Towel", value, flags=re.I)
    value = re.sub(r"\bT\s+oilet\b", "Toilet", value, flags=re.I)
    value = re.sub(r"\bT\s+echnology\b", "Technology", value, flags=re.I)
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def extract_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return clean_text("\n".join(page.extract_text() or "" for page in reader.pages))


def split_sections(text: str) -> dict[str, str]:
    matches = []
    for name in SECTION_NAMES:
        pos = text.find(name)
        if pos >= 0:
            matches.append((pos, name))
    matches.sort()

    sections = {}
    for index, (pos, name) in enumerate(matches):
        end = matches[index + 1][0] if index + 1 < len(matches) else len(text)
        sections[name] = text[pos + len(name):end].strip()
    return sections


def parse_header(text: str) -> tuple[str, datetime, str, str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    try:
        report_index = lines.index("Maintenance Report")
        property_name = lines[report_index + 1]
    except (ValueError, IndexError) as exc:
        raise ValueError("Could not detect the property name in the PDF header.") from exc

    date_match = DATE_RE.search(text)
    if not date_match:
        raise ValueError("Could not detect the inspection date in the PDF header.")

    inspection_date = datetime.strptime(
        f"{date_match.group(2)} {date_match.group(3)} {date_match.group(4)}",
        "%B %d %Y",
    )

    address = ""
    date_line_index = next((i for i, line in enumerate(lines) if DATE_RE.search(line)), None)
    if date_line_index is not None:
        address_parts = []
        for line in lines[date_line_index + 1:date_line_index + 3]:
            if "Grand Welcome" not in line and line.lower() != "none":
                address_parts.append(line)
        address = ", ".join(address_parts)

    slug = re.sub(r"[^A-Z0-9]+", "-", property_name.upper()).strip("-")
    inspection_id = f"{slug}-{inspection_date:%Y-%m-%d}"
    return property_name, inspection_date, address, inspection_id


def parse_marked_section(section: str):
    lines = [line.strip() for line in section.splitlines() if line.strip()]
    records = []
    i = 0
    markers = ("Count ", "Yes / No ", "Checklist", "Text", "Photo", "Rating ")

    while i < len(lines):
        line = lines[i]

        if line.startswith("Count "):
            raw = line.split(" ", 1)[1].strip()
            answer = None if raw.lower() == "none" else int(raw)
            prompt = lines[i + 1] if i + 1 < len(lines) else ""
            records.append(("count", prompt, answer))
            i += 2
            continue

        if line.startswith("Yes / No "):
            answer = line.split("Yes / No ", 1)[1].strip()
            prompt = lines[i + 1] if i + 1 < len(lines) else ""
            records.append(("yesno", prompt, answer))
            i += 2
            continue

        if line == "Checklist":
            prompt = lines[i + 1] if i + 1 < len(lines) else ""
            records.append(("checklist", prompt, "Yes"))
            i += 2
            continue

        if line == "Photo":
            prompt = lines[i + 1] if i + 1 < len(lines) else ""
            records.append(("photo", prompt, "Photo included"))
            i += 2
            continue

        if line == "Text":
            prompt = lines[i + 1] if i + 1 < len(lines) else ""
            i += 2
            answer_lines = []
            while i < len(lines) and not any(lines[i].startswith(marker) for marker in markers):
                answer_lines.append(lines[i])
                i += 1
            records.append(("text", prompt, " ".join(answer_lines).strip()))
            continue

        if line.startswith("Rating "):
            stars = line.split("Rating ", 1)[1]
            rating = stars.count("★")
            prompt = lines[i + 1] if i + 1 < len(lines) else ""
            records.append(("rating", prompt, rating))
            i += 2
            continue

        i += 1

    return records


def classify_requirement(item: str) -> tuple[str, Optional[int]]:
    lower = item.lower()
    if "1.5x occupancy" in lower:
        return "1.5x Occupancy", None
    if "plunger" in lower or "toilet brush" in lower:
        return "Bathrooms", None
    if "hair dryer" in lower:
        return "Manual", 2
    if "space heater" in lower:
        return "None", None
    return "Manual", None


def priority_for(question: str) -> str:
    lower = question.lower()
    if any(term in lower for term in [
        "smoke detector", "co detector", "fire extinguisher", "smart lock", "fireplace"
    ]):
        return "Critical"
    if any(term in lower for term in [
        "wifi", "tv", "garage opener", "parking pass", "hoa amenity",
        "exterior lighting", "mattress protector", "pillow protector"
    ]):
        return "High"
    if any(term in lower for term in ["snow shovel", "ice melt"]):
        return "Seasonal"
    return "Standard"


def timeframe_for(prompt: str) -> str:
    match = re.search(r"\(([^)]+)\)", prompt)
    return match.group(1) if match else "Review"


def parse_report(path: Path, source: str) -> ParsedReport:
    text = extract_pdf_text(path)
    property_name, inspection_date, address, inspection_id = parse_header(text)

    report = ParsedReport(
        property_name=property_name,
        inspection_date=inspection_date,
        address=address,
        source=source,
        inspection_id=inspection_id,
    )

    for category, body in split_sections(text).items():
        for kind, prompt, answer in parse_marked_section(body):
            if kind == "count":
                requirement_type, manual_required = classify_requirement(prompt)
                report.inventory.append([
                    inspection_id, property_name, inspection_date, category, prompt,
                    requirement_type, manual_required, None, None, None, "",
                    None, answer, None, None, "", ""
                ])

            elif kind in ("yesno", "checklist"):
                if category == "Kitchen Inventory" and "1.5x occupancy" in prompt.lower():
                    requirement_type, manual_required = classify_requirement(prompt)
                    report.inventory.append([
                        inspection_id, property_name, inspection_date, category, prompt,
                        requirement_type, manual_required, None, None, None, answer,
                        None, None, None, None, "", ""
                    ])
                else:
                    report.conditions.append([
                        inspection_id, property_name, inspection_date, category, prompt,
                        answer, priority_for(prompt), "", "", "", None, "No"
                    ])

            elif kind == "text":
                report.feedback.append([
                    inspection_id, property_name, inspection_date, category, prompt,
                    answer, timeframe_for(prompt), "", "", "", None, ""
                ])

            elif kind == "rating":
                report.ratings.append([
                    inspection_id, property_name, inspection_date, prompt, answer, ""
                ])

    return report


def download_pdf(url: str) -> Path:
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    if not response.content.startswith(b"%PDF"):
        raise ValueError("The URL did not return a PDF file.")
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    handle.write(response.content)
    handle.close()
    return Path(handle.name)


def get_property_dimensions(workbook, property_name: str):
    if "Properties" not in workbook.sheetnames:
        return None, None, None

    sheet = workbook["Properties"]
    target = property_name.strip().lower()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0] and str(row[0]).strip().lower() == target:
            bedrooms = row[1]
            bathrooms = row[2]
            sleeps = row[3]
            return bedrooms, bathrooms, sleeps

    return None, None, None


def existing_inspection_ids(workbook) -> set[str]:
    if "Inspections" not in workbook.sheetnames:
        return set()
    sheet = workbook["Inspections"]
    return {
        str(cell.value).strip()
        for cell in sheet["A"][1:]
        if cell.value is not None and str(cell.value).strip()
    }


def append_rows(sheet, rows: list[list]):
    for row in rows:
        sheet.append(row)


def populate_inventory_formulas(sheet, start_row: int, end_row: int):
    for row in range(start_row, end_row + 1):
        requirement_type = sheet[f"F{row}"].value
        manual_required = sheet[f"G{row}"].value
        actual = sheet[f"M{row}"].value
        compliance = sheet[f"K{row}"].value

        if requirement_type == "1.5x Occupancy":
            sheet[f"L{row}"] = f"=ROUNDUP(J{row}*1.5,0)"
        elif requirement_type == "Bathrooms":
            sheet[f"L{row}"] = f"=I{row}"
        elif requirement_type == "Bedrooms":
            sheet[f"L{row}"] = f"=H{row}"
        elif requirement_type == "Manual" and manual_required not in (None, ""):
            sheet[f"L{row}"] = manual_required

        has_required = (
            requirement_type in ("1.5x Occupancy", "Bathrooms", "Bedrooms")
            or (requirement_type == "Manual" and manual_required not in (None, ""))
        )

        if has_required and actual not in (None, ""):
            sheet[f"N{row}"] = f"=M{row}-L{row}"
            sheet[f"O{row}"] = f"=MAX(0,-N{row})"

        if compliance == "No":
            sheet[f"P{row}"] = "ACTION — QUANTITY TBD"
        elif compliance == "Yes":
            sheet[f"P{row}"] = "MEETS PAR"
        elif has_required and actual not in (None, ""):
            sheet[f"P{row}"] = f'=IF(O{row}>0,"REPLACE / REPLENISH "&O{row},"MEETS REQUIREMENT")'
        elif requirement_type == "None":
            sheet[f"P{row}"] = "INFORMATIONAL"
        else:
            sheet[f"P{row}"] = "SET REQUIRED LEVEL"


def import_report(pdf_path: Path, source: str, workbook_path: Path, force: bool = False):
    report = parse_report(pdf_path, source)
    workbook = load_workbook(workbook_path)

    if not force and report.inspection_id in existing_inspection_ids(workbook):
        return {
            "status": "Duplicate",
            "property": report.property_name,
            "date": report.inspection_date,
            "inspection_id": report.inspection_id,
            "message": "Inspection already exists and was skipped.",
        }

    required_sheets = [
        "Inspections", "Inventory", "Condition Checks",
        "Feedback & Actions", "Ratings", "Import Queue"
    ]
    missing = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing:
        raise ValueError("Workbook is missing required sheets: " + ", ".join(missing))

    bedrooms, bathrooms, sleeps = get_property_dimensions(workbook, report.property_name)

    for row in report.inventory:
        row[7] = bedrooms
        row[8] = bathrooms
        row[9] = sleeps

    inspections = workbook["Inspections"]
    inspections.append([
        report.inspection_id,
        report.property_name,
        report.inspection_date,
        "",
        report.address,
        source,
        f"Q{((report.inspection_date.month - 1) // 3) + 1} {report.inspection_date.year}",
        "",
    ])

    inventory = workbook["Inventory"]
    inventory_start = inventory.max_row + 1
    append_rows(inventory, report.inventory)
    inventory_end = inventory.max_row
    if inventory_end >= inventory_start:
        populate_inventory_formulas(inventory, inventory_start, inventory_end)

    conditions = workbook["Condition Checks"]
    condition_start = conditions.max_row + 1
    append_rows(conditions, report.conditions)
    for row in range(condition_start, conditions.max_row + 1):
        conditions[f"H{row}"] = (
            f'=IF(F{row}="No","ACTION REQUIRED",'
            f'IF(F{row}="Needs Attention","REVIEW",'
            f'IF(F{row}="Yes","PASS",IF(F{row}="N/A","N/A",""))))'
        )

    feedback = workbook["Feedback & Actions"]
    feedback_start = feedback.max_row + 1
    append_rows(feedback, report.feedback)
    for row in range(feedback_start, feedback.max_row + 1):
        feedback[f"I{row}"] = (
            f'=IF(OR(F{row}="",F{row}="0"),"NO ACTION",'
            f'IF(L{row}<>"","COMPLETE",'
            f'IF(H{row}="Approved","APPROVED — SCHEDULE",'
            f'IF(H{row}="Declined","DECLINED","OPEN"))))'
        )

    ratings = workbook["Ratings"]
    rating_start = ratings.max_row + 1
    append_rows(ratings, report.ratings)
    for row in range(rating_start, ratings.max_row + 1):
        ratings[f"F{row}"] = (
            f'=IF(E{row}=5,"Excellent",IF(E{row}=4,"Good",'
            f'IF(E{row}=3,"Fair",IF(E{row}=2,"Needs Improvement","Poor"))))'
        )

    queue = workbook["Import Queue"]
    queue.append([
        "Web App",
        source,
        report.property_name,
        report.inspection_date,
        report.inspection_id,
        "Imported",
        datetime.now(),
        (
            f"Added {len(report.inventory)} inventory rows, "
            f"{len(report.conditions)} condition rows, "
            f"{len(report.feedback)} feedback rows, "
            f"and {len(report.ratings)} ratings."
        ),
    ])

    workbook.save(workbook_path)

    return {
        "status": "Imported",
        "property": report.property_name,
        "date": report.inspection_date,
        "inspection_id": report.inspection_id,
        "message": (
            f"Imported {len(report.inventory)} inventory, "
            f"{len(report.conditions)} condition, "
            f"{len(report.feedback)} feedback, and "
            f"{len(report.ratings)} rating records."
        ),
    }
